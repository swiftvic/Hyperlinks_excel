# generate_hyperlinks.py
# May 17, 2019
import openpyxl
import re                                                                                     # Regular Expressions library
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

'''
Add hyperlinks to an excel file from a list. Hyperlinks are generated dynamically from the list of part numbers in column A.

'''

# Settings and variables
wb1_filepath = "S:\TEST\Quotes\Sangoma\Sangoma Test Quote Matrix.xlsx"
hyperlink = '=HYPERLINK("\\\mt.local\corp\DATA\Quote_Cost\Customer Quote information\Sangoma\Ireland Takeover\\'

wb1_ws = "PartNumList"

def open_files(wb1):
    '''
    Opens workbook1 and workbook2 file paths and assigns to workbook1 and workbook2
    '''
    wb1_path = wb1

    workbook1 = openpyxl.load_workbook(wb1_path)

    return workbook1

def stats(ws):
    '''
    Pass in the worksheet and will print stats of the max column and row of sheet.
    '''
    max_row = ws.max_row
    max_col = ws.max_column

    print("There are " + str(max_row) + " rows and " + str(max_col) + " columns in " + str(ws) + ".")

def delete_col(ws, *colnum):
    '''
    Deletes passed in columns from passed in worksheet
    *argv allows you to passin multiple columns that needs to be deleted
    ws = worksheet, *colnum = column number(s) (can add more column numbers that will be deleted)
    '''
    for col in colnum:
        ws.delete_cols(col)
        print("Column " + str(col) + " has been deleted.")

if __name__ == "__main__":
    wb1 = open_files(wb1_filepath)
    ws = wb1[wb1_ws]
    stats(ws)
    count = 0       #flag to count number of hyperlinks

    for r in range(3, ws.max_row):
#        print (hyperlink + ws.cell(r, 1).value + '", "' + ws.cell(r, 1).value + '")')
        ws.cell(r, 2).value = (hyperlink + ws.cell(r, 1).value + '", "' + ws.cell(r, 1).value + '")')
        count +=1

    print(str(count) + " Hyperlinks added.")

    delete_col(ws, 1)

    wb1.save("S:\TEST\Quotes\Sangoma\Sangoma Test Quote Matrix-Linked.xlsx")

